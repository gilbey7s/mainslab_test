import openpyxl
from rest_framework import  viewsets, status
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.shortcuts import get_object_or_404
from pathlib import Path

from .serializers import *
from .models import *
from .filters import BillsFilterSet
from .utils import detector, classifier


validate_date = {
    'suffix':['.xlsx', ],
    'name_file':['bills', 'client_org'],
}

@api_view(['POST'])
def download_excel_view(request):
    path = Path(request.data['file'])
    suffix_file = str(Path(path.suffix))
    name_file = str(Path(path.stem))
    if suffix_file not in validate_date['suffix']:
        suffixs =  validate_date['suffix']
        message = f"Неверное расширение файла. Подходящие форматы {suffixs}"
        return Response({'error': message})
    if name_file not in validate_date['name_file']:
        names_file =  validate_date['name_file']
        message = (
                "Неизвестное имя файла. "
                f"Обрабатываемый файл должен иметь имя - {names_file}"
        )
        return Response({'error': message})
    book = openpyxl.open(request.data['file'])
    if name_file == 'bills': 
        sheet = book.active
        for row in range(2, sheet.max_row + 1):
            client_name = sheet[row][0].value
            client_org = sheet[row][1].value
            order_number= sheet[row][2].value
            sum = sheet[row][3].value
            date = sheet[row][4].value
            service = sheet[row][5].value
            fraud_score = detector(service)
            if fraud_score >= 0.9:
                clent_org_obj = get_object_or_404(ClientOrganization, name=client_org)
                clent_org_obj.fraud_weight =+1
                clent_org_obj.save()
            service_dict=classifier(service)
            service_class = list(service_dict.keys())[0]
            service_name = list(service_dict.values())[0]
            client_name_obj = get_object_or_404(Client, name=client_name)
            client_org_obj = get_object_or_404(ClientOrganization, name=client_org)
            Bill(client_name=client_name_obj,
                    client_org=client_org_obj,
                    order_number=order_number,
                    sum=sum, date=date,
                    service=service,
                    fraud_score=fraud_score,
                    service_class=service_class,
                    service_name=service_name,
                    ).save()
            
    if name_file == 'client_org': 
        sheet = book.worksheets[0]
        for row in range(2, sheet.max_row + 1):
            name = sheet[row][0].value
            Client(name=name).save()
        sheet_next = book.worksheets[1]
        for row in range(2, sheet_next.max_row + 1):
            client_name = sheet_next[row][0].value
            name = sheet_next[row][1].value
            address = str(sheet_next[row][2].value)
            if address != ' ' != '-':
                address ='Адрес: '+ address
            client_name_obj = get_object_or_404(Client, name=client_name)
            ClientOrganization(client_name=client_name_obj, name=name, address=address).save()
    return Response({'message': 1}, status=status.HTTP_201_CREATED)


class ClientsViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Client.objects.all()
    serializer_class = ClientSerializer


class BillsViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Bill.objects.all()
    serializer_class = BillSerializers
    filter_backends = (DjangoFilterBackend,)
    filterset_class = BillsFilterSet
